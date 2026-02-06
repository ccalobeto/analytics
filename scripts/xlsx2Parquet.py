import pandas as pd
from openpyxl import load_workbook
import re
import unicodedata

def to_snake_case(col):
    # Normalize accents (AÑO -> ANO)
    col = unicodedata.normalize("NFKD", str(col)).encode("ascii", "ignore").decode()

    # Lowercase
    col = col.lower()

    # Replace non-alphanumeric with underscore
    col = re.sub(r"[^a-z0-9]+", "_", col)

    # Remove leading/trailing underscores
    col = col.strip("_")

    return col

# Inputs
xlsx_file = "data/Cronologico total.xlsx"
sheet_name = "CRONO"   # or sheet index like 0
csv_out = "./user_files/cronologico.csv"
parquet_out = "./user_files/cronologico.parquet"
skiprows = 12  # Number of rows to reach data

# Load workbook in safe mode (avoids pivot cache parsing)
wb = load_workbook(xlsx_file, read_only=True, data_only=True)
ws = wb[sheet_name]

rows = []
for row in ws.iter_rows(values_only=True):
    rows.append(row)

# Build DataFrame
df = pd.DataFrame(rows[skiprows:], columns=rows[skiprows - 1]) 

# Apply normalization
df.columns = [to_snake_case(c) for c in df.columns]

# remove the columns when all the values are None
df = df.dropna(axis=1, how='all')
df = df.iloc[:, :-1]  # remove last empty column if exists

print(df.columns.tolist())
# Filter rows where 'fch_anomes_ped' is numeric
df = df[pd.to_numeric(df["fch_anomes_ped"], errors="coerce").notna()]

# invalids -> NaT
df["fch_entrg"] = pd.to_datetime(df["fch_entrg"], errors="coerce")  

# Write outputs
df.to_csv(csv_out, index=False)
df.to_parquet(parquet_out, index=False)
